from datetime import datetime
import os
from pathlib import Path
import openpyxl as opxl


class ExcelHandler:

    def __init__(self, folder_path: str, file_name: str):
        self.file_path = Path(folder_path) / file_name
        self.wb = self._load_workbook()

    def _load_workbook(self):
        """Tenta carregar a pasta de trabalho de forma segura."""
        if not self.file_path.exists():
            print(f"❌ Erro: O arquivo '{self.file_path}' não foi encontrado.")
            return None
        try:
            return opxl.load_workbook(self.file_path)
        except Exception as e:
            print(f"❌ Erro ao abrir o arquivo Excel: {e}")
            return None

    @property
    def is_valid(self) -> bool:
        """Retorna se o arquivo foi carregado com sucesso."""
        return self.wb is not None

    def _is_date(self, value) -> bool:
        """Verifica se um valor é um objeto datetime ou pode ser convertido em data."""
        if value is None:
            return False
        if isinstance(value, datetime):
            return True

        val_str = str(value).strip()
        formats_to_try = [
            "%m/%Y",
            "%d/%m/%Y",
            "%Y-%m-%d",
            "%d-%m-%Y",
            "%m-%Y",
            "%d/%m/%y",
        ]

        for fmt in formats_to_try:
            try:
                datetime.strptime(val_str, fmt)
                return True
            except ValueError:
                continue

        return False

    def get_sheet_names(self) -> list[str]:
        """Retorna os nomes das abas disponíveis no arquivo."""
        return self.wb.sheetnames if self.is_valid else []

    def copy_sheet(self, origin: str, destination: str) -> bool:
        """Copia o conteúdo de uma aba para outra no mesmo arquivo."""
        if not self.is_valid:
            return False

        if origin not in self.wb.sheetnames:
            print(f"❌ A aba de origem '{origin}' não existe.")
            return False

        if destination in self.wb.sheetnames:
            del self.wb[destination]

        ws_origin = self.wb[origin]
        ws_dest = self.wb.copy_worksheet(ws_origin)
        ws_dest.title = destination

        print(f"✔️ Conteúdo de '{origin}' copiado para '{destination}'.")
        return True

    def clean_sheet(self, sheet_name: str) -> bool:
        """Desmescla todas as células e remove a quebra automática de texto."""
        if not self.is_valid or sheet_name not in self.wb.sheetnames:
            print(f"❌ Aba '{sheet_name}' não encontrada para tratamento.")
            return False

        ws = self.wb[sheet_name]

        # 1. Desmesclar células
        merged_ranges = list(ws.merged_cells.ranges)
        for cell_range in merged_ranges:
            ws.unmerge_cells(str(cell_range))

        # 2. Remover quebra automática de texto (wrap_text = False)
        for row in ws.iter_rows():
            for cell in row:
                if cell.alignment and cell.alignment.wrap_text:
                    ws.cell(row=cell.row, column=cell.column).alignment = (
                        cell.alignment.copy(wrap_text=False)
                    )

        print(f"🧹 Aba '{sheet_name}' desmesclada e sem quebra de texto.")
        return True

    def trim_rows_by_anchors(
        self,
        sheet_name: str,
        header_text: str = "CÓDIGO",
        footer_text: str = "RESUMO DO BALANCETE",
        col_idx: int = 1,
    ) -> bool:
        """Localiza dinamicamente o cabeçalho e o rodapé na coluna especificada (padrão: Coluna A)

        e remove tudo o que estiver fora desse intervalo.
        """
        if not self.is_valid or sheet_name not in self.wb.sheetnames:
            return False

        ws = self.wb[sheet_name]
        header_row = None
        footer_row = None

        # 1. Varre a Coluna A identificando as linhas das âncoras
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col_idx).value
            if val is not None:
                cell_str = str(val).strip().upper()

                if header_text.upper() in cell_str and header_row is None:
                    header_row = row

                if footer_text.upper() in cell_str and footer_row is None:
                    footer_row = row

        # 2. Apaga o rodapé primeiro (preserva os índices superiores)
        if footer_row:
            rows_to_delete_footer = (ws.max_row - footer_row) + 1
            ws.delete_rows(footer_row, rows_to_delete_footer)
            print(
                f"✂️ Rodapé encontrado ('{footer_text}' na linha {footer_row}). Removidas {rows_to_delete_footer} linhas do fim."
            )
        else:
            print(
                f"ℹ️ Texto de rodapé '{footer_text}' não encontrado na Coluna A."
            )

        # 3. Apaga o topo (tudo antes do cabeçalho)
        if header_row and header_row > 1:
            rows_to_delete_header = header_row - 1
            ws.delete_rows(1, rows_to_delete_header)
            print(
                f"✂️ Cabeçalho encontrado ('{header_text}' na linha {header_row}). Removidas as primeiras {rows_to_delete_header} linhas."
            )
        elif header_row == 1:
            print(f"ℹ️ O cabeçalho '{header_text}' já está na linha 1.")
        else:
            print(
                f"⚠️ Texto de cabeçalho '{header_text}' não foi encontrado na Coluna A."
            )

        return True

    def delete_empty_columns(self, sheet_name: str) -> bool:
        """Identifica e remove colunas que não possuem nenhum dado útil."""
        if not self.is_valid or sheet_name not in self.wb.sheetnames:
            return False

        ws = self.wb[sheet_name]

        # Itera de trás para frente (da última coluna até a coluna 1)
        for col_idx in range(ws.max_column, 0, -1):
            is_empty = True

            for row in range(1, ws.max_row + 1):
                val = ws.cell(row=row, column=col_idx).value

                if val is not None and str(val).strip() != "":
                    is_empty = False
                    break

            if is_empty:
                ws.delete_cols(col_idx)

        print(f"🗑️ Colunas totalmente vazias removidas na aba '{sheet_name}'.")
        return True

    def filter_rows_by_last_column_date(self, sheet_name: str) -> bool:
        """Verifica a primeira linha da última coluna com valor.

        Se NÃO for data, remove as linhas onde essa coluna estiver vazia.
        """
        if not self.is_valid or sheet_name not in self.wb.sheetnames:
            return False

        ws = self.wb[sheet_name]
        last_col_idx = ws.max_column

        if last_col_idx < 1:
            return False

        header_value = ws.cell(row=1, column=last_col_idx).value

        # Se for data, não faz alterações
        if self._is_date(header_value):
            print(
                f"ℹ️ A última coluna ({last_col_idx}) é uma data ('{header_value}'). Nenhuma linha foi removida."
            )
            return True

        print(
            f"⚠️ A última coluna ({last_col_idx}) NÃO é data ('{header_value}'). Filtrando linhas vazias..."
        )

        rows_deleted = 0
        # Deleta de baixo para cima preservando o cabeçalho (linha 1)
        for r in range(ws.max_row, 1, -1):
            val = ws.cell(row=r, column=last_col_idx).value
            if val is None or str(val).strip() == "":
                ws.delete_rows(r)
                rows_deleted += 1

        print(
            f"✂️ Foram excluídas {rows_deleted} linhas onde a última coluna estava vazia."
        )
        return True

    def save(self) -> None:
        """Salva as alterações no arquivo original."""
        if self.is_valid:
            self.wb.save(self.file_path)
            print(f"💾 Arquivo salvo com sucesso em: '{self.file_path}'")

    def process_and_prepare(self, origin: str, destination: str) -> None:
        """Esteira linear de execução:

        1. Copia a aba de origem
        2. Desmescla células e remove quebras
        3. Delimita por 'CÓDIGO' e 'RESUMO DO BALANCETE'
        4. Exclui colunas vazias
        5. Avalia/filtra a última coluna caso não seja data
        6. Salva as alterações
        """
        if self.copy_sheet(origin, destination):
            self.clean_sheet(destination)
            self.trim_rows_by_anchors(
                destination,
                header_text="CÓDIGO",
                footer_text="RESUMO DO BALANCETE",
            )
            self.delete_empty_columns(destination)
            self.filter_rows_by_last_column_date(destination)
            self.save()


# ==========================================
# EXECUÇÃO DO SCRIPT
# ==========================================

if __name__ == "__main__":
    folder_path = input("O caminho da pasta:\n").strip()
    file_name = input("Nome do arquivo (ex: dados.xlsx):\n").strip()

    excel = ExcelHandler(folder_path, file_name)

    if excel.is_valid:
        print("\n✅ Arquivo encontrado!")
        print(f"Abas disponíveis: {excel.get_sheet_names()}\n")

        origin_sheet = input("Aba de origem:\n").strip()
        destination_sheet = input("Aba de destino:\n").strip()

        excel.process_and_prepare(origin_sheet, destination_sheet)